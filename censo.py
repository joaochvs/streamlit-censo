# ── Imports ───────────────────────────────────────────────────────────────────

import re
import unicodedata
import datetime
from io import BytesIO

import pandas as pd
import streamlit as st
from rapidfuzz import process, fuzz
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Constantes ────────────────────────────────────────────────────────────────

COLUNAS_PROD = [
    "Visitados", "Pesquisas", "Recusas", "Produtivos",
    "Abandonados", "Ausentes", "Em Construção"
]

SCORE_MATCH   = 95
SCORE_REVISAR = 75


# ── Funções utilitárias ───────────────────────────────────────────────────────

def normalizar(nome: str) -> str:
    if pd.isna(nome):
        return ""
    nome = str(nome).lower()
    nome = unicodedata.normalize("NFKD", nome)
    nome = "".join(c for c in nome if not unicodedata.combining(c))
    nome = re.sub(r"[^a-z0-9 ]", "", nome)
    return " ".join(nome.split())


def quebrar_nome(nome: str) -> str:
    partes = nome.split()
    if len(partes) > 3:
        return " ".join(partes[:3]) + "<br/>" + " ".join(partes[3:])
    return nome


# WRatio já combina token_set, partial e outros — mas sozinho ele pode dar
# score alto só porque um SOBRENOME comum bate (ex.: "Martins", "Oliveira"),
# mesmo que o primeiro nome seja outra pessoa completamente diferente. Por
# isso adicionamos duas checagens de sanidade abaixo antes de aceitar o
# score do WRatio.
def calcular_score(nome1: str, nome2: str, **kwargs) -> float:
    score_wratio = fuzz.WRatio(nome1, nome2)

    tokens1 = nome1.split()
    tokens2 = nome2.split()
    if not tokens1 or not tokens2:
        return score_wratio

    # 1) Primeiro nome muito diferente → provavelmente é outra pessoa,
    #    mesmo que o sobrenome bata (ex.: "Adriane Martins" x
    #    "Sonia Aparecida Santos Martins Camargo").
    score_primeiro_nome = fuzz.ratio(tokens1[0], tokens2[0])
    if score_primeiro_nome < 85:
        score_wratio = min(score_wratio, score_primeiro_nome)

    # 2) Um nome com 1 palavra só comparado a outro com número diferente de
    #    palavras é evidência fraca demais para um match automático
    #    (ex.: "Ana" x "Joana Maciel", "Ana" x "Ana Julia Ferreira").
    #    Um nome curto pode "casar" por dentro de outro (partial_ratio),
    #    então não confiamos nesses casos sem revisão manual.
    if min(len(tokens1), len(tokens2)) == 1 and len(tokens1) != len(tokens2):
        score_wratio = min(score_wratio, SCORE_REVISAR - 1)

    return score_wratio


# Preposições/conectores comuns em nomes em português: ficam em minúsculo,
# exceto quando são a primeira palavra do nome.
CONECTORES_NOME = {"de", "da", "do", "das", "dos", "e"}


def capitalizar_nome(nome: str) -> str:
    if pd.isna(nome):
        return nome
    nome = str(nome).strip()
    if not nome:
        return nome

    palavras   = nome.lower().split()
    resultado  = []
    for i, palavra in enumerate(palavras):
        if palavra in CONECTORES_NOME and i != 0:
            resultado.append(palavra)
        else:
            resultado.append(palavra[:1].upper() + palavra[1:])
    return " ".join(resultado)


# ── Carregamento (cache por conteúdo do arquivo, não por data) ────────────────

@st.cache_data(show_spinner="Carregando censo...")
def carregar_censo(
    arquivos_bytes: list[bytes],   # bytes em vez de file objects → cache estável
    data_inicio: datetime.date,
    data_fim: datetime.date,
) -> pd.DataFrame:
    partes = []
    for b in arquivos_bytes:
        df = pd.read_excel(BytesIO(b))
        # FIX: pd.to_datetime() sem format infere o padrão a partir das
        # primeiras linhas (ex.: %m/%d/%Y) e quebra em datas como 30/06/2026,
        # já que não existe mês 30. Como as datas aqui são no padrão
        # brasileiro (dia primeiro), usamos dayfirst=True + format="mixed"
        # para que cada valor seja interpretado individualmente e
        # corretamente, tanto se vier como texto quanto como data nativa
        # do Excel.
        df["Data do registro"] = pd.to_datetime(
            df["Data do registro"], format="mixed", dayfirst=True
        ).dt.date
        df["Nome do Agente"] = df["Nome do Agente"].apply(capitalizar_nome)
        partes.append(df)
    df_total = pd.concat(partes, ignore_index=True).drop_duplicates()
    return df_total[
        (df_total["Data do registro"] >= data_inicio) &
        (df_total["Data do registro"] <= data_fim)
    ].copy()


@st.cache_data(show_spinner="Carregando CSV...")
def carregar_csv(
    arquivo_bytes: bytes,
    data_inicio: datetime.date,
    data_fim: datetime.date,
) -> pd.DataFrame:
    df = pd.read_csv(BytesIO(arquivo_bytes))
    # FIX: mesmo ajuste aplicado aqui — evita o ValueError
    # "time data ... doesn't match format %m/%d/%Y" quando o dia do mês
    # é maior que 12 (ex.: 30/06/2026).
    df["Data"] = pd.to_datetime(
        df["Data"], format="mixed", dayfirst=True
    ).dt.date
    df["Agente"] = df["Agente"].apply(capitalizar_nome)
    df = df[
        (df["Data"] >= data_inicio) &
        (df["Data"] <= data_fim)
    ].copy()
    cols_existentes = [c for c in COLUNAS_PROD if c in df.columns]
    df["produtividade_total"] = df[cols_existentes].sum(axis=1)
    return df.reset_index(drop=True)


# ── Agregação ─────────────────────────────────────────────────────────────────

def agregar_censo(df_censo: pd.DataFrame) -> pd.DataFrame:
    df = (
        df_censo.groupby("Nome do Agente")
        .size()
        .reset_index(name="produtividade_censo")
        .sort_values("produtividade_censo", ascending=False)
    )
    df["nome_normalizado"] = df["Nome do Agente"].apply(normalizar)
    return df


def agregar_csv(df_csv: pd.DataFrame) -> pd.DataFrame:
    df = (
        df_csv.groupby("Agente")["produtividade_total"]
        .sum()
        .reset_index(name="produtividade_informada")
        .sort_values("produtividade_informada", ascending=False)
    )
    df["nome_normalizado"] = df["Agente"].apply(normalizar)
    return df


# ── Comparação fuzzy ──────────────────────────────────────────────────────────

def comparar(df_csv_prod: pd.DataFrame, df_censo_prod: pd.DataFrame) -> pd.DataFrame:
    nomes_censo      = df_censo_prod["nome_normalizado"].tolist()
    resultado        = []
    nomes_csv_matched = set()

    for _, linha in df_csv_prod.iterrows():
        melhor = process.extractOne(
            linha["nome_normalizado"], nomes_censo, scorer=calcular_score
        )

        if melhor is None or melhor[1] < SCORE_REVISAR:
            resultado.append({
                "Agente Informado":        linha["Agente"],
                "Agente Sistema":          "Não encontrado",
                "Similaridade":            round(melhor[1], 1) if melhor else 0,
                "Produtividade Informada": linha["produtividade_informada"],
                "Produtividade Sistema":   0,
                "Diferença":               linha["produtividade_informada"],
                "Status":                  "❌ Sem correspondência",
            })
            continue

        nome_match, score, _ = melhor
        agente_sis = df_censo_prod[df_censo_prod["nome_normalizado"] == nome_match].iloc[0]
        nomes_csv_matched.add(nome_match)

        diferenca = linha["produtividade_informada"] - agente_sis["produtividade_censo"]
        status    = "✅ Match" if score >= SCORE_MATCH else "⚠ Revisar"

        resultado.append({
            "Agente Informado":        linha["Agente"],
            "Agente Sistema":          agente_sis["Nome do Agente"],
            "Similaridade":            round(score, 1),
            "Produtividade Informada": linha["produtividade_informada"],
            "Produtividade Sistema":   agente_sis["produtividade_censo"],
            "Diferença":               diferenca,
            "Status":                  status,
        })

    # agentes só no sistema
    for _, ag in df_censo_prod[~df_censo_prod["nome_normalizado"].isin(nomes_csv_matched)].iterrows():
        resultado.append({
            "Agente Informado":        "Não encontrado",
            "Agente Sistema":          ag["Nome do Agente"],
            "Similaridade":            0,
            "Produtividade Informada": 0,
            "Produtividade Sistema":   ag["produtividade_censo"],
            "Diferença":               -ag["produtividade_censo"],
            "Status":                  "📋 Só no sistema",
        })

    return pd.DataFrame(resultado)


# ── Geração de PDF ────────────────────────────────────────────────────────────

def gerar_pdf(
    df_resultado: pd.DataFrame,
    data_inicio: datetime.date,
    data_fim: datetime.date,
) -> BytesIO:
    buffer    = BytesIO()
    doc       = SimpleDocTemplate(buffer, pagesize=letter)
    styles    = getSampleStyleSheet()
    elementos = []

    meses = [
        "janeiro", "fevereiro", "março", "abril", "maio", "junho",
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
    ]

    def por_extenso(d):
        return f"{d.day} de {meses[d.month - 1]} de {d.year}"

    periodo = (
        por_extenso(data_inicio)
        if data_inicio == data_fim
        else f"{por_extenso(data_inicio)} a {por_extenso(data_fim)}"
    )

    elementos.append(Paragraph(
        "<b><font size=14>Relatório de Divergência de Produtividade</font></b>",
        styles["Title"],
    ))
    elementos.append(Paragraph(" ", styles["Normal"]))
    elementos.append(Paragraph(
        f"<b>Período de referência:</b><br/>{periodo}", styles["Normal"]
    ))
    elementos.append(Paragraph(" ", styles["Normal"]))

    colunas_pdf = [
        "Agente Informado", "Agente Sistema",
        "Produtividade Informada", "Produtividade Sistema", "Diferença"
    ]
    cabecalhos  = ["Agente (Formulário)", "Agente (Sistema)", "Formulário", "Sistema", "Diferença"]

    df_t = df_resultado.sort_values("Diferença", ascending=False)[colunas_pdf].copy()
    df_t.columns = cabecalhos
    df_t["Agente (Formulário)"] = df_t["Agente (Formulário)"].apply(quebrar_nome)
    for col in ["Formulário", "Sistema", "Diferença"]:
        df_t[col] = df_t[col].round(0).astype(int)

    data_table = [cabecalhos]
    for _, row in df_t.iterrows():
        data_table.append([
            Paragraph(str(row["Agente (Formulário)"]), styles["Normal"]),
            Paragraph(str(row["Agente (Sistema)"]),    styles["Normal"]),
            str(row["Formulário"]),
            str(row["Sistema"]),
            str(row["Diferença"]),
        ])

    tabela = Table(data_table, colWidths=[140, 140, 60, 60, 60])
    style  = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("ALIGN",      (2, 1), (-1, -1), "CENTER"),
    ])

    for i, row in enumerate(df_t.itertuples(), start=1):
        cor = colors.blue if row.Diferença < 0 else (colors.red if row.Diferença > 0 else colors.black)
        style.add("TEXTCOLOR", (-1, i), (-1, i), cor)
        style.add("FONTNAME",  (-1, i), (-1, i), "Helvetica-Bold")

    tabela.setStyle(style)
    elementos.append(tabela)
    doc.build(elementos)
    buffer.seek(0)
    return buffer


# ── Geração de Excel ──────────────────────────────────────────────────────────
# Mesmo layout do PDF (Agente Formulário | Agente Sistema | Formulário | Sistema | Diferença)
# + coluna Data → permite filtrar por dia no Excel

def gerar_excel(
    df_resultado: pd.DataFrame,
    df_csv_raw: pd.DataFrame,
    data_inicio: datetime.date,
    data_fim: datetime.date,
) -> BytesIO:

    # ── helpers de estilo ─────────────────────────────────────────────────────
    DARK  = "1F3864"
    BEGE  = "FFF8DC"
    PAR   = "EBF3FB"
    TOTAL = "FCE4D6"
    WHITE = "FFFFFF"

    thin  = Side(style="thin", color="AAAAAA")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fl(cor): return PatternFill("solid", fgColor=cor)
    def fn(bold=False, color="000000", size=9): return Font(bold=bold, color=color, name="Arial", size=size)
    def al(h="center", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ── construir tabela de dados por agente x data ───────────────────────────
    # Para cada agente, para cada dia do período: prod informada e prod sistema
    mapa_sistema = (
        df_resultado[df_resultado["Agente Sistema"] != "Não encontrado"]
        .set_index("Agente Informado")["Agente Sistema"]
        .to_dict()
    )
    mapa_prod_sistema = (
        df_resultado[df_resultado["Agente Sistema"] != "Não encontrado"]
        .set_index("Agente Informado")["Produtividade Sistema"]
        .to_dict()
    )

    # prod informada por agente por dia
    prod_dia = (
        df_csv_raw.groupby(["Agente", "Data"])["produtividade_total"]
        .sum()
        .reset_index()
    )
    prod_dia.columns = ["Agente", "Data", "Prod_Informada"]

    # prod sistema por agente: vem do censo, sem info de dia → dividir igualmente pelos dias com registro
    # estratégia: mostrar prod sistema total na linha de resumo e deixar dias com a informada
    datas_com_registro = sorted(prod_dia["Data"].unique())

    linhas = []
    for agente in prod_dia["Agente"].unique():
        dias_agente = prod_dia[prod_dia["Agente"] == agente].set_index("Data")["Prod_Informada"]
        prod_sis_total = mapa_prod_sistema.get(agente, 0)
        nome_sis       = mapa_sistema.get(agente, "Não encontrado")

        for data in datas_com_registro:
            prod_inf = int(dias_agente.get(data, 0))
            linhas.append({
                "Data":                    data,
                "Agente (Formulário)":     agente,
                "Agente (Sistema)":        nome_sis,
                "Formulário (dia)":        prod_inf,
                "Sistema (total período)": int(prod_sis_total),
            })

    # Agentes que existem só no sistema (têm produtividade no censo, mas
    # não foram informados no formulário) também entram aqui, com
    # "Não encontrado" no Formulário e 0 de produtividade informada — assim
    # eles não ficam de fora do Excel, só aparecendo hoje no resumo.
    so_no_sistema = df_resultado[df_resultado["Agente Informado"] == "Não encontrado"]
    for _, linha_sis in so_no_sistema.iterrows():
        nome_sis       = linha_sis["Agente Sistema"]
        prod_sis_total = int(linha_sis["Produtividade Sistema"])

        for data in datas_com_registro:
            linhas.append({
                "Data":                    data,
                "Agente (Formulário)":     "Não encontrado",
                "Agente (Sistema)":        nome_sis,
                "Formulário (dia)":        0,
                "Sistema (total período)": prod_sis_total,
            })

    df_excel = pd.DataFrame(linhas).sort_values(["Data", "Agente (Formulário)"])

    # ── montar workbook ───────────────────────────────────────────────────────
    wb = Workbook()

    # ────────────────────────────────────────────────────────────────────────
    # ABA 1 — Por Dia (igual ao PDF mas com Data)
    # ────────────────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Por Dia"

    cabecalhos = [
        "Data",
        "Agente (Formulário)",
        "Agente (Sistema)",
        "Formulário (dia)",
        "Sistema (período)",
        "Diferença",
    ]
    larguras = [13, 28, 28, 16, 16, 12]

    for i, (cab, larg) in enumerate(zip(cabecalhos, larguras), 1):
        ws1.column_dimensions[get_column_letter(i)].width = larg

    # cabeçalho
    ws1.row_dimensions[1].height = 20
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cabecalhos))
    c = ws1.cell(1, 1)
    c.value     = f"Divergência de Produtividade por Dia  —  {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
    c.font      = fn(bold=True, color=WHITE, size=11)
    c.fill      = fl(DARK)
    c.alignment = al()

    ws1.row_dimensions[2].height = 16
    for col, cab in enumerate(cabecalhos, 1):
        c = ws1.cell(2, col)
        c.value     = cab
        c.font      = fn(bold=True, color=WHITE, size=9)
        c.fill      = fl(DARK)
        c.alignment = al()
        c.border    = borda

    ws1.freeze_panes = "A3"

    # dados
    for idx, (_, row) in enumerate(df_excel.iterrows(), start=3):
        ws1.row_dimensions[idx].height = 14
        cor = PAR if idx % 2 == 0 else WHITE

        prod_inf = int(row["Formulário (dia)"])
        prod_sis = int(row["Sistema (total período)"])
        dif_val  = prod_inf - prod_sis

        vals = [
            row["Data"].strftime("%d/%m/%Y"),
            row["Agente (Formulário)"],
            row["Agente (Sistema)"],
            prod_inf,
            prod_sis,
            dif_val,
        ]

        for col, val in enumerate(vals, 1):
            c = ws1.cell(idx, col)
            c.value     = val
            c.font      = fn(size=9)
            c.fill      = fl(cor)
            c.border    = borda
            c.alignment = al(h="left" if col <= 3 else "center")

        # colorir diferença
        cor_dif = "C00000" if dif_val > 0 else ("1F497D" if dif_val < 0 else "000000")
        ws1.cell(idx, 6).font = fn(bold=True, color=cor_dif, size=9)

    # ────────────────────────────────────────────────────────────────────────
    # ABA 2 — Resumo (igual ao PDF: totais por agente)
    # ────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo Geral")

    df_res = df_resultado.sort_values("Diferença", ascending=False)
    cabs2  = ["Agente (Formulário)", "Agente (Sistema)", "Formulário", "Sistema", "Diferença", "Status"]
    largs2 = [28, 28, 12, 12, 12, 18]

    for i, larg in enumerate(largs2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = larg

    ws2.row_dimensions[1].height = 20
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cabs2))
    c = ws2.cell(1, 1)
    c.value     = f"Resumo Geral  —  {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
    c.font      = fn(bold=True, color=WHITE, size=11)
    c.fill      = fl(DARK)
    c.alignment = al()

    ws2.row_dimensions[2].height = 16
    for col, cab in enumerate(cabs2, 1):
        c = ws2.cell(2, col)
        c.value     = cab
        c.font      = fn(bold=True, color=WHITE, size=9)
        c.fill      = fl(DARK)
        c.alignment = al()
        c.border    = borda

    ws2.freeze_panes = "A3"

    for idx, (_, row) in enumerate(df_res.iterrows(), start=3):
        ws2.row_dimensions[idx].height = 14
        cor = PAR if idx % 2 == 0 else WHITE

        dif_val = int(row["Diferença"])
        vals = [
            row["Agente Informado"],
            row["Agente Sistema"],
            int(row["Produtividade Informada"]),
            int(row["Produtividade Sistema"]),
            dif_val,
            row["Status"],
        ]

        for col, val in enumerate(vals, 1):
            c = ws2.cell(idx, col)
            c.value     = val
            c.font      = fn(size=9)
            c.fill      = fl(cor)
            c.border    = borda
            c.alignment = al(h="left" if col <= 2 else "center")

        cor_dif = "C00000" if dif_val > 0 else ("1F497D" if dif_val < 0 else "000000")
        ws2.cell(idx, 5).font = fn(bold=True, color=cor_dif, size=9)

    # linha de totais
    row_t = 3 + len(df_res)
    ws2.row_dimensions[row_t].height = 16

    ws2.cell(row_t, 1).fill   = fl(TOTAL)
    ws2.cell(row_t, 1).border = borda
    c = ws2.cell(row_t, 2)
    c.value     = "TOTAL GERAL"
    c.font      = fn(bold=True, size=9)
    c.fill      = fl(TOTAL)
    c.alignment = al(h="right")
    c.border    = borda

    for col, letra in [(3, "C"), (4, "D"), (5, "E")]:
        c = ws2.cell(row_t, col)
        c.value         = f"=SUM({letra}3:{letra}{row_t-1})"
        c.font          = fn(bold=True, size=9)
        c.fill          = fl(TOTAL)
        c.alignment     = al()
        c.border        = borda
        c.number_format = "#,##0"

    ws2.cell(row_t, 6).fill   = fl(TOTAL)
    ws2.cell(row_t, 6).border = borda

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Interface Streamlit ───────────────────────────────────────────────────────

st.set_page_config(layout="centered")
st.title("Divergência de Produtividade — Agentes")

# ── Período ───────────────────────────────────────────────────────────────────

periodo = st.date_input(
    "Selecione o período:",
    value=(datetime.date.today(), datetime.date.today()),
    format="DD/MM/YYYY",
)

if len(periodo) < 2:
    st.info("Selecione também a data final do período.")
    st.stop()

data_inicio, data_fim = periodo

# ── Upload ────────────────────────────────────────────────────────────────────

col1, col2 = st.columns(2)

with col1:
    arquivos_censo = st.file_uploader(
        "📊 Base do sistema / censo (XLSX)",
        type=["xlsx"],
        accept_multiple_files=True,
        key="censo",
    )

with col2:
    arquivo_csv = st.file_uploader(
        "📄 Base de produtividade (CSV)",
        type=["csv"],
        key="csv",
    )

# ── Processamento ─────────────────────────────────────────────────────────────

if arquivo_csv and arquivos_censo:
    st.caption(f"📂 {len(arquivos_censo)} arquivo(s) de censo carregado(s)")

    # lê bytes uma vez → cache estável mesmo se o objeto file mudar
    bytes_censo = [f.read() for f in arquivos_censo]
    bytes_csv   = arquivo_csv.read()

    df_censo = carregar_censo(bytes_censo, data_inicio, data_fim)
    df_csv   = carregar_csv(bytes_csv, data_inicio, data_fim)

    if df_censo.empty:
        st.warning("Nenhum registro encontrado no censo para o período selecionado.")
        st.stop()

    if df_csv.empty:
        st.warning("Nenhum registro encontrado no CSV para o período selecionado.")
        st.stop()

    df_censo_prod = agregar_censo(df_censo)
    df_csv_prod   = agregar_csv(df_csv)

    with st.spinner("Realizando comparação fuzzy..."):
        df_resultado = comparar(df_csv_prod, df_censo_prod)

    df_divergencia = df_resultado[
        (df_resultado["Diferença"] > 0) &
        (df_resultado["Status"] != "❌ Sem correspondência")
    ].sort_values("Diferença", ascending=False)

    # ── Métricas ──────────────────────────────────────────────────────────────
    m1, m2, m3 = st.columns(3)
    m1.metric("Agentes no formulário", len(df_csv_prod))
    m2.metric("Agentes no sistema",    len(df_censo_prod))
    m3.metric("Com divergência",       len(df_divergencia))

    st.divider()

    # ── Tabelas ───────────────────────────────────────────────────────────────
    st.subheader("📋 Produtividade Informada")
    st.dataframe(df_csv_prod.drop(columns="nome_normalizado"), use_container_width=True)

    st.subheader("🚨 Quem informou mais do que fez")
    st.dataframe(df_divergencia, use_container_width=True)

    st.subheader("🔍 Comparativo Final")
    st.dataframe(
        df_resultado.sort_values(["Status", "Diferença"], ascending=[True, False]),
        use_container_width=True,
    )

    # ── Downloads ─────────────────────────────────────────────────────────────
    sufixo = (
        str(data_inicio)
        if data_inicio == data_fim
        else f"{data_inicio}_a_{data_fim}"
    )

    col_pdf, col_xls = st.columns(2)

    with col_pdf:
        pdf = gerar_pdf(df_resultado, data_inicio, data_fim)
        st.download_button(
            label="📄 Baixar Relatório PDF",
            data=pdf,
            file_name=f"relatorio_divergencia_{sufixo}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

    with col_xls:
        excel = gerar_excel(df_resultado, df_csv, data_inicio, data_fim)
        st.download_button(
            label="📊 Baixar Diário Excel",
            data=excel,
            file_name=f"diario_produtividade_{sufixo}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )